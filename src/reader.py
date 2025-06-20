
import openpyxl
import tkinter as tk
from tkinter import filedialog

class ExcelReader:
    def __init__(self):
        self.file_path = ""
        self.workbook = None
        self.sheet = None

    def select_file(self):
        # Open file dialog to select Excel file
        root = tk.Tk()
        root.withdraw()  # Hide the root window
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xlsm")])
        
        if file_path:
            self.file_path = file_path
            self.load_workbook()
        else:
            print("Dosyayı unuttun galiba, neyse.")
            return False
        return True

    def load_workbook(self):
        if self.file_path:
            try:
                self.workbook = openpyxl.load_workbook(self.file_path)
                self.sheet = self.workbook.active  # Use the active sheet by default
                print(f"Seçtiğin dosya bu muydu: {self.file_path}\n")
            except Exception as e:
                print(f"Bugun git yarın gel, problem cikti: {e}")
                self.workbook = None
                self.sheet = None

    def resize_cells(self):
        if not self.select_file():
            return
        
        print("Enter the column letter to resize (e.g., A, B, C):")
        col = input().strip().upper()

        print("Enter the new width for this column:")
        width = float(input().strip())

        if self.sheet and col and width:
            self.sheet.column_dimensions[col].width = width
            print(f"Column {col} resized to width {width}.")
            self.save_changes()

    def check_condition(self):
        if not self.select_file():
            return

        print("Enter the condition you want to check (e.g., >10, <5):")
        condition = input().strip()

        print("Enter the column letter to apply the condition to:")
        col = input().strip().upper()

        if self.sheet and condition and col:
            condition_func = self.create_condition_func(condition)
            self.apply_condition(col, condition_func)

    def create_condition_func(self, condition):
        # A simple condition check (extendable)
        if condition.startswith(">"):
            value = float(condition[1:])
            return lambda cell: cell.value > value
        elif condition.startswith("<"):
            value = float(condition[1:])
            return lambda cell: cell.value < value
        else:
            print("Invalid condition format.")
            return None

    def apply_condition(self, col, condition_func):
        if not condition_func:
            return

        # Apply the condition to all rows in the column
        for row in self.sheet.iter_rows(min_col=openpyxl.utils.column_index_from_string(col), 
                                        max_col=openpyxl.utils.column_index_from_string(col)):
            cell = row[0]
            if condition_func(cell):
                print(f"Cell {cell.coordinate} meets the condition ({cell.value}).")
        
        self.save_changes()

    def save_changes(self):
        if self.workbook:
            self.workbook.save(self.file_path)
            print(f"Yazdım bunları bak, unutmam: {self.file_path}.")
